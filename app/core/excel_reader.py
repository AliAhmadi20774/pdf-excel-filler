from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


class ExcelReader:
    def __init__(self, path: str | Path):
        self.path = Path(path)

    def read_columns(self) -> list[str]:
        workbook = load_workbook(self.path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            return [str(cell).strip() for cell in headers if cell is not None and str(cell).strip()]
        finally:
            workbook.close()

    def read_rows(self) -> list[dict]:
        workbook = load_workbook(self.path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return []
            headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
            results: list[dict] = []
            for row in rows[1:]:
                record = {}
                for index, header in enumerate(headers):
                    if not header:
                        continue
                    value = row[index] if index < len(row) else None
                    record[header] = value
                if any(value is not None and value != "" for value in record.values()):
                    results.append(record)
            return results
        finally:
            workbook.close()
